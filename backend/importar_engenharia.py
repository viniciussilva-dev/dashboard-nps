# === SCRIPT: importar_engenharia.py ===
# Arquivo: backend/importar_engenharia.py
#
# O que este script faz:
# 1. APAGA as respostas antigas da tabela respostas_engenharia
# 2. Importa as 10 respostas do Excel com conversão correta:
#    - Palavras (Excelente, Bom...) → números na escala 1-10
#    - Isso porque o novo Typebot envia notas de 1 a 10
#
# Conversão usada:
#   Excelente → 10
#   Bom       → 8
#   Regular   → 6
#   Ruim      → 4
#   Péssimo   → 2
#
# Como executar (dentro da pasta backend/):
#   python importar_engenharia.py

import openpyxl
import mysql.connector
import os
from dotenv import load_dotenv

# === CARREGA AS VARIÁVEIS DO ARQUIVO .env ===
load_dotenv()

# === CONFIGURAÇÃO DO BANCO ===
config = {
    'host':     os.getenv('DB_HOST', '127.0.0.1'),
    'port':     int(os.getenv('DB_PORT', 3306)),
    'user':     os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'canalsolar_nps'),
    'charset':  'utf8mb4',
}

# === FUNÇÃO: Converte palavra para nota na escala 1-10 ===
# O Excel tem os valores como texto: "Excelente", "Bom", etc.
# Convertemos para a mesma escala que o novo Typebot usa (1-10)
# Exemplos:
#   "Excelente" → 10
#   "Bom"       → 8
#   "Regular"   → 6
#   "Ruim"      → 4
#   "Péssimo"   → 2
#   None        → None (campo vazio, salva como nulo)
def palavra_para_nota(v):
    if v is None or str(v).strip() == '':
        return None
    # Se já vier como número, retorna direto
    try:
        return int(float(v))
    except:
        pass
    mapa = {
        'excelente': 10,
        'bom':        8,
        'regular':    6,
        'ruim':       4,
        'péssimo':    2,
        'pessimo':    2,
    }
    return mapa.get(str(v).lower().strip())

# === CAMINHO DO EXCEL ===
ARQUIVO_EXCEL = 'Pesquisa_de_Satisfação_Engenharia.xlsx'
ABA = 'Respostas ao formulário 2'

print('=' * 60)
print('📥 Importação — Respostas de Engenharia (escala 1-10)')
print('=' * 60)

# === ABRE O EXCEL ===
try:
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
    ws = wb[ABA]
    print(f'✅ Excel aberto com sucesso')
    print(f'📊 Aba: "{ABA}"')
    print(f'📋 Total de respostas encontradas: {ws.max_row - 1}')
    print()
except FileNotFoundError:
    print(f'❌ Arquivo "{ARQUIVO_EXCEL}" não encontrado!')
    print(f'   Coloque o Excel na pasta backend/ e tente novamente.')
    exit(1)

# === CONECTA AO BANCO ===
try:
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()
    print(f'✅ Conectado ao MySQL em {config["host"]}')
    print(f'🗄️  Banco: {config["database"]}')
    print()
except Exception as e:
    print(f'❌ Erro ao conectar ao banco: {e}')
    exit(1)

# === APAGA AS RESPOSTAS ANTIGAS ===
# Limpa a tabela para evitar duplicatas ao reimportar
print('🗑️  Apagando respostas antigas...')
cursor.execute('DELETE FROM respostas_engenharia')
deletadas = cursor.rowcount
print(f'   {deletadas} respostas removidas')
print()

# === SQL DE INSERÇÃO ===
sql = """
    INSERT INTO respostas_engenharia
        (nome, email, empresa,
         avaliacao_agilidade,
         avaliacao_conhecimento_tecnico,
         avaliacao_qualidade,
         avaliacao_pontualidade,
         avaliacao_satisfacao,
         indicaria_amigo,
         data_resposta,
         ano)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

# === LOOP DE IMPORTAÇÃO ===
sucesso = 0
erros   = 0

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):

    # Posições das colunas na aba "Respostas ao formulário 2":
    # [0] Carimbo de data/hora
    # [1] Seu nome
    # [2] Seu E-mail
    # [3] Nome da empresa
    # [4] Agilidade no atendimento
    # [5] Conhecimento técnico dos funcionários
    # [6] Qualidade do trabalho entregue
    # [7] Pontualidade das entregas
    # [8] Satisfação com o serviço
    # [9] Indicaria para um amigo (NPS 0-10)
    # [10] Elogio ou reclamação (feedback)

    data_original = row[0]
    nome          = row[1]
    email         = row[2]
    empresa       = row[3]
    agilidade     = row[4]
    conh_tec      = row[5]
    qualidade     = row[6]
    pontualidade  = row[7]
    satisfacao    = row[8]
    nps_raw       = row[9]

    # NPS é obrigatório
    if nps_raw is None:
        print(f'  ⚠️  Linha {i} ignorada — campo NPS vazio')
        erros += 1
        continue

    # Formata a data
    if data_original:
        data_str = data_original.strftime('%Y-%m-%d %H:%M:%S')
        ano      = data_original.year
    else:
        from datetime import datetime
        agora    = datetime.now()
        data_str = agora.strftime('%Y-%m-%d %H:%M:%S')
        ano      = agora.year

    # Converte palavras para números na escala 1-10
    nota_agilidade    = palavra_para_nota(agilidade)
    nota_conh_tec     = palavra_para_nota(conh_tec)
    nota_qualidade    = palavra_para_nota(qualidade)
    nota_pontualidade = palavra_para_nota(pontualidade)
    nota_satisfacao   = palavra_para_nota(satisfacao)

    valores = (
        str(nome).strip()    if nome    else None,
        str(email).strip()   if email   else None,
        str(empresa).strip() if empresa else None,
        nota_agilidade,
        nota_conh_tec,
        nota_qualidade,
        nota_pontualidade,
        nota_satisfacao,
        int(nps_raw),
        data_str,
        ano,
    )

    try:
        cursor.execute(sql, valores)
        print(f'  ✅ Linha {i:02d} | {str(empresa)[:25]:<25} | NPS: {int(nps_raw):>2} | Agilidade: {nota_agilidade}/10')
        sucesso += 1
    except Exception as e:
        print(f'  ❌ Linha {i} — Erro: {e}')
        erros += 1

# === CONFIRMA NO BANCO ===
conn.commit()
cursor.close()
conn.close()

print()
print('=' * 60)
print(f'✅ Importação concluída!')
print(f'   Inseridas com sucesso: {sucesso} respostas')
if erros > 0:
    print(f'   Erros/ignoradas:       {erros} linhas')
print('=' * 60)